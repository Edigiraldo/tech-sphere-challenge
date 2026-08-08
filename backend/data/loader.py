"""Read-only loaders for the synthetic dataset XLSX files.

All functions return immutable dataclass instances.  The source XLSX/PDF files
are never modified.

Two conversation-loading entry points exist:

* ``load_conversations()`` — runtime-safe (``label_ground_truth`` always
  ``None``).
* ``load_conversations_for_evaluation()`` — evaluation-only; populates
  ``label_ground_truth`` on every turn.

The evaluation-only function is deliberately omitted from ``__init__.py``'s
public re-exports so that CI / static-analysis gates can flag accidental use in
production code.
"""

from __future__ import annotations

import datetime
import json
import logging
import os
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

from .models import Conversation, ConversationTurn, Patient, Trajectory

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Locate the dataset root (workspace-relative, usable from any cwd)
# ---------------------------------------------------------------------------

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
_DATASET_PATH = _PROJECT_ROOT / "dataset"


def get_dataset_path() -> Path:
    """Absolute path to the ``dataset/`` directory."""
    return _DATASET_PATH


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _parse_date(value) -> Optional[datetime.date]:
    """Parse an Excel serial date, datetime, or date string into a date."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
        try:
            return datetime.date.fromisoformat(value.strip())
        except (ValueError, TypeError):
            pass
    return None


def _parse_date_required(value, patient_id: str) -> datetime.date:
    """Parse a required date field, raising on failure.

    Unlike ``_parse_date`` this does **not** silently fall back to today's
    date — a missing or unparseable ``fecha_cirugia`` would corrupt post-op
    day calculations and lead to incorrect clinical decisions.
    """
    result = _parse_date(value)
    if result is not None:
        return result
    logger.error(
        "Cannot parse fecha_cirugia=%r for patient %s — raising ValueError",
        value, patient_id,
    )
    raise ValueError(
        f"fecha_cirugia is missing or unparseable for patient {patient_id}: "
        f"{value!r}"
    )


def _parse_json_list(value) -> List[str]:
    """Parse a JSON-list string (used by ``comorbilidades`` and
    ``adaptation_fields``) into a list of strings."""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item) for item in value]
    if isinstance(value, str):
        value = value.strip()
        if value in ("", "[]", "null", "None"):
            return []
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list):
                return [str(item) for item in parsed]
        except (json.JSONDecodeError, TypeError):
            pass
    return []


def _coerce_int(value, *, field: str = "", context: str = "") -> int:
    """Coerce a value to int, defaulting to 0.

    Logs a warning when the value cannot be parsed, which may indicate
    data corruption in the source spreadsheet.
    """
    try:
        return int(value)
    except (TypeError, ValueError):
        logger.warning(
            "Coercing non-int %s=%r to 0%s%s — possible data corruption",
            field, value,
            " in " if context else "", context,
        )
        return 0


def _coerce_float(value, *, field: str = "", context: str = "") -> float:
    """Coerce a value to float, defaulting to 0.0.

    Logs a warning when the value cannot be parsed, which may indicate
    data corruption in the source spreadsheet.
    """
    try:
        return float(value)
    except (TypeError, ValueError):
        logger.warning(
            "Coercing non-float %s=%r to 0.0%s%s — possible data corruption",
            field, value,
            " in " if context else "", context,
        )
        return 0.0


def _coerce_str(value) -> str:
    """Coerce a value to str, defaulting to ''."""
    if value is None:
        return ""
    return str(value)


def _coerce_bool(value) -> bool:
    """Coerce a value to bool."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in ("true", "1", "yes", "si", "sí")
    return False


# ---------------------------------------------------------------------------
# Patient loading (merged clinical + demographic)
# ---------------------------------------------------------------------------

def load_patients() -> Dict[str, Patient]:
    """Load all 40 synthetic patients, merging clinical and demographic profiles.

    Returns a dict keyed by ``paciente_id``.
    """
    # -- clinical profiles --
    clin_path = _DATASET_PATH / "perfiles_clinicos_pacientes_silver_contest.xlsx"
    clinical: Dict[str, dict] = {}
    wb_clin = openpyxl.load_workbook(clin_path, read_only=True, data_only=True)
    ws = wb_clin[wb_clin.sheetnames[0]]
    headers_clin = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers_clin, row))
        pid = str(record.get("paciente_id", ""))
        if not pid:
            continue
        clinical[pid] = record
    wb_clin.close()

    # -- demographic profiles --
    demo_path = _DATASET_PATH / "perfiles_pacientes_co.xlsx"
    demographic: Dict[str, dict] = {}
    wb_demo = openpyxl.load_workbook(demo_path, read_only=True, data_only=True)
    ws2 = wb_demo[wb_demo.sheetnames[0]]
    headers_demo = [cell.value for cell in next(ws2.iter_rows(min_row=1, max_row=1))]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers_demo, row))
        pid = str(record.get("paciente_id", ""))
        if not pid:
            continue
        demographic[pid] = record
    wb_demo.close()

    # -- merge --
    result: Dict[str, Patient] = {}
    for pid, clin in clinical.items():
        demo = demographic.get(pid, {})
        patient = Patient(
            paciente_id=pid,
            bundle_id=_coerce_str(clin.get("bundle_id")),
            synthea_runtime=_coerce_str(clin.get("synthea_runtime")),
            modulo_synthea=_coerce_str(clin.get("modulo_synthea")),
            procedimiento=_coerce_str(clin.get("procedimiento")),
            fecha_cirugia=_parse_date_required(
                clin.get("fecha_cirugia"), pid
            ),
            edad=_coerce_int(clin.get("edad"), field="edad", context=f"patient {pid}"),
            genero=_coerce_str(clin.get("genero")),
            comorbilidades=_parse_json_list(clin.get("comorbilidades")),
            complicacion_encounter=_coerce_bool(clin.get("complicacion_encounter")),
            nombre_completo=_coerce_str(demo.get("nombre_completo")),
            direccion=_coerce_str(demo.get("direccion")),
            ciudad=_coerce_str(demo.get("ciudad")),
            departamento=_coerce_str(demo.get("departamento")),
            documento_cc=_coerce_str(demo.get("documento_cc")),
            eps=_coerce_str(demo.get("eps")),
            source_country=_coerce_str(demo.get("source_country")),
            adapted_country=_coerce_str(demo.get("adapted_country")),
            adaptation_fields=_parse_json_list(
                demo.get("adaptation_fields")
            ),
        )
        result[pid] = patient
    return result


# ---------------------------------------------------------------------------
# Trajectory loading
# ---------------------------------------------------------------------------

def load_trajectories() -> Dict[str, List[Trajectory]]:
    """Load all post-operative trajectories grouped by ``paciente_id``.

    Each patient may have multiple trajectories (one per post-op day).
    """
    path = _DATASET_PATH / "trayectorias_postop_silver.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    result: Dict[str, List[Trajectory]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        pid = str(record.get("paciente_id", ""))
        if not pid:
            continue
        trajectory = Trajectory(
            trayectoria_id=_coerce_str(record.get("trayectoria_id")),
            paciente_id=pid,
            dia_postop=_coerce_int(record.get("dia_postop"), field="dia_postop", context=f"patient {pid}"),
            arquetipo_trayectoria=_coerce_str(
                record.get("arquetipo_trayectoria")
            ),
            dolor_nrs=_coerce_int(record.get("dolor_nrs"), field="dolor_nrs", context=f"patient {pid}"),
            fiebre_c=_coerce_float(record.get("fiebre_c"), field="fiebre_c", context=f"patient {pid}"),
            movilidad=_coerce_str(record.get("movilidad")),
            herida=_coerce_str(record.get("herida")),
            apetito=_coerce_str(record.get("apetito")),
            sueno=_coerce_str(record.get("sueno")),
            seed=_coerce_int(record.get("seed"), field="seed", context=f"patient {pid}"),
        )
        result.setdefault(pid, []).append(trajectory)
    wb.close()
    return result


# ---------------------------------------------------------------------------
# Conversation loading (runtime-safe — no label_ground_truth)
# ---------------------------------------------------------------------------

def _load_conversation_rows(
    include_label: bool,
) -> Dict[str, List[ConversationTurn]]:
    """Shared row loader.  ``include_label`` controls whether the
    ``label_ground_truth`` column is read from the source."""
    path = _DATASET_PATH / "dataset_final.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    by_caso: Dict[str, List[ConversationTurn]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        caso_id = str(record.get("caso_id", ""))
        if not caso_id:
            continue

        turn = ConversationTurn(
            dialogo_id=_coerce_str(record.get("dialogo_id")),
            caso_id=caso_id,
            paciente_id=_coerce_str(record.get("paciente_id")),
            dia_postop=_coerce_int(record.get("dia_postop"), field="dia_postop", context=f"caso {caso_id}"),
            turno_idx=_coerce_int(record.get("turno_idx"), field="turno_idx", context=f"caso {caso_id}"),
            hablante=_coerce_str(record.get("hablante")),
            texto=_coerce_str(record.get("texto")),
            label_ground_truth=(
                _coerce_str(record.get("label_ground_truth"))
                if include_label
                else None
            ),
            estilo_paciente=_coerce_str(record.get("estilo_paciente")),
            modelo_paciente=_coerce_str(record.get("modelo_paciente")),
            modelo_agente=_coerce_str(record.get("modelo_agente")),
            capa=_coerce_str(record.get("capa")),
        )
        by_caso.setdefault(caso_id, []).append(turn)
    wb.close()

    # Sort turns within each conversation
    for turns in by_caso.values():
        turns.sort(key=lambda t: t.turno_idx)

    return by_caso


def _validate_conversation_homogeneity(
    caso_id: str, turns: List[ConversationTurn],
) -> tuple[str, int]:
    """Verify all turns share the same ``paciente_id`` and ``dia_postop``.

    Returns the validated (paciente_id, dia_postop) pair.
    Raises ``ValueError`` on inconsistency — silently trusting the first turn
    would mask data corruption.
    """
    if not turns:
        raise ValueError(f"Conversation {caso_id} has no turns")
    pid = turns[0].paciente_id
    dia = turns[0].dia_postop
    for turn in turns[1:]:
        if turn.paciente_id != pid:
            raise ValueError(
                f"Conversation {caso_id}: paciente_id mismatch: "
                f"turn {turn.turno_idx} has {turn.paciente_id!r}, "
                f"expected {pid!r}"
            )
        if turn.dia_postop != dia:
            raise ValueError(
                f"Conversation {caso_id}: dia_postop mismatch: "
                f"turn {turn.turno_idx} has {turn.dia_postop}, "
                f"expected {dia}"
            )
    return pid, dia


def load_conversations() -> List[Conversation]:
    """Load all synthetic conversations grouped by ``caso_id``.

    **label_ground_truth is always None.**  This is the runtime-safe entry point.
    Use ``load_conversations_for_evaluation()`` when you need the labels for
    offline evaluation or testing.

    Raises ``ValueError`` if any conversation contains turns with inconsistent
    ``paciente_id`` or ``dia_postop`` values.
    """
    by_caso = _load_conversation_rows(include_label=False)
    result: List[Conversation] = []
    for caso_id, turns in sorted(by_caso.items()):
        pid, dia = _validate_conversation_homogeneity(caso_id, turns)
        result.append(
            Conversation(
                caso_id=caso_id,
                paciente_id=pid,
                dia_postop=dia,
                turns=turns,
            )
        )
    return result


# ---------------------------------------------------------------------------
# Conversation loading (evaluation-only — includes label_ground_truth)
# ---------------------------------------------------------------------------

def load_conversations_for_evaluation() -> List[Conversation]:
    """Load conversations **with** ``label_ground_truth`` populated.

    This function exists exclusively for offline evaluation, testing, and
    metrics.  Production/runtime modules must NOT call it — the field is not
    used during live conversations and must not influence escalation,
    summarisation, or any clinical decision path.

    Raises ``ValueError`` if any conversation contains turns with inconsistent
    ``paciente_id`` or ``dia_postop`` values.
    """
    by_caso = _load_conversation_rows(include_label=True)
    result: List[Conversation] = []
    for caso_id, turns in sorted(by_caso.items()):
        pid, dia = _validate_conversation_homogeneity(caso_id, turns)
        result.append(
            Conversation(
                caso_id=caso_id,
                paciente_id=pid,
                dia_postop=dia,
                turns=turns,
            )
        )
    return result
